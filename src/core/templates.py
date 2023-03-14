import contextlib
import logging
import pathlib
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core import models


def generate_template_file(file_path: pathlib.Path) -> None:
    workbook = openpyxl.Workbook(write_only=True)

    with contextlib.closing(workbook):
        prices_worksheet: Worksheet = workbook.create_sheet('Цены')
        prices_worksheet.append(('nm ID', 'Новая цена'))

        stocks_worksheet: Worksheet = workbook.create_sheet('Остатки')
        stocks_worksheet.append(('ID склада', 'sku', 'Количество остатков'))

        workbook.save(file_path)


def generate_prices_report_file(
        *,
        file_path: pathlib.Path,
        categories_prices: Iterable[models.CategoryPrices],
) -> None:
    workbook = openpyxl.Workbook(write_only=True)

    with contextlib.closing(workbook):
        worksheet: Worksheet = workbook.create_sheet('Цены')
        worksheet.append(('Категория товара', 'nm ID', 'Новая цена', 'Скидка'))

        for category_prices in categories_prices:

            for nomenclature_price in category_prices.nomenclature_prices:
                worksheet.append((
                    category_prices.category_name,
                    nomenclature_price.nomenclature_id,
                    nomenclature_price.price,
                    nomenclature_price.discount,
                ))

        workbook.save(file_path)


class Template:

    def __init__(self, file_path: str | pathlib.Path):
        self.__file_path = file_path
        self.__workbook = openpyxl.Workbook()
        self.__prices_worksheet: Worksheet = self.__workbook.active
        self.__prices_worksheet.title = 'Цены'
        self.__stocks_worksheet: Worksheet = self.__workbook.create_sheet('Остатки')
        self.__write_headers()
        self.__is_saved = False

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        self.__workbook.close()

    def save(self):
        if not self.__is_saved:
            self.__workbook.save(self.__file_path)
            logging.debug(f'Template file was created in {self.__file_path}')
            self.__is_saved = True

    def __write_headers(self):
        self.__prices_worksheet.append(('nm ID', 'Цена', 'Скидка (не указывать)'))
        self.__stocks_worksheet.append(('ID склада', 'sku', 'Количество остатков'))

    def add_prices_rows(
            self,
            category_to_nomenclature_prices: dict[str, Iterable[models.NomenclaturePrice]],
    ) -> None:
        for category, nomenclature_price in category_to_nomenclature_prices.items():
            self.__prices_worksheet.append(
                (nomenclature_price.nomenclature_id, nomenclature_price.price, nomenclature_price.discount)
            )

    def add_stocks_rows(self, warehouses_stocks: Iterable[models.WarehouseStocks]) -> None:
        for warehouse_stocks in warehouses_stocks:
            for stock in warehouse_stocks.stocks:
                self.__stocks_worksheet.append((warehouse_stocks.warehouse_id, stock.sku, stock.amount))
