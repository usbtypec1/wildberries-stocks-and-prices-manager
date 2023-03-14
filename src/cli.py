import collections
import logging
import pathlib
from typing import DefaultDict

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from rich.console import Console
from rich.progress import track
from rich.prompt import Prompt, Confirm

from core import exceptions
from core import models
from core.helpers import grouper
from core.parser import WorkbookParser
from core.services.http_clients import closing_wildberries_api_http_client
from core.services.wildberries_api import WildberriesAPIService
from core.templates import generate_stocks_report_file, generate_template_file, generate_prices_report_file
from core.validators import validate_api_key


def upload_prices(console: Console, api_key: str, file_path: str | pathlib.Path):
    workbook = openpyxl.load_workbook(file_path, read_only=True)

    parser = WorkbookParser(workbook)
    try:
        nomenclature_prices = parser.get_prices()
    except exceptions.WorksheetMissingError as error:
        console.log(f'В файле отсутствует страница "{error.worksheet_name}"')
        return
    console.log('Цены считаны из файла')

    with closing_wildberries_api_http_client(api_key=api_key) as http_client:
        wildberries_api_service = WildberriesAPIService(http_client)

        for nomenclature_prices_group in track(
                sequence=tuple(grouper(nomenclature_prices, n=1000)),
                description='Загрузка цен...',
        ):
            wildberries_api_service.update_prices(nomenclature_prices_group)


def upload_stocks(console: Console, api_key: str, file_path: str | pathlib.Path):
    workbook = openpyxl.load_workbook(file_path, read_only=True)

    parser = WorkbookParser(workbook)
    try:
        warehouses_stocks = parser.get_stocks()
    except exceptions.WorksheetMissingError as error:
        console.log(f'В файле отсутствует страница "{error.worksheet_name}"')
        return
    console.log('Остатки считаны из файла')

    with closing_wildberries_api_http_client(api_key=api_key) as http_client:
        wildberries_api_service = WildberriesAPIService(http_client)

        for warehouse_stocks in warehouses_stocks:
            console.log(f'Загрузка остатков в складе №{warehouse_stocks.warehouse_id}')

            for stocks_group in track(
                    sequence=tuple(grouper(warehouse_stocks.stocks, n=100)),
                    description='Загрузка остатков...',
            ):
                try:
                    wildberries_api_service.update_stocks(warehouse_stocks.warehouse_id, stocks_group)
                except Exception:
                    logging.exception('Error while updating stocks')


def download_stocks(console: Console, api_key: str, file_path: str | pathlib.Path):
    warehouse_id_to_stocks_by_skus: DefaultDict[int, list[models.StocksBySku]] = collections.defaultdict(list)
    skus: set[str] = set()

    with closing_wildberries_api_http_client(api_key=api_key) as http_client:
        wildberries_api_service = WildberriesAPIService(http_client)

        warehouses = wildberries_api_service.get_warehouses()
        console.log('Информация о складах загружена')
        for nomenclatures in track(
                sequence=wildberries_api_service.get_nomenclatures(),
                description='Загрузка всех skus...',
                console=console,
        ):
            skus |= {
                sku
                for nomenclature in nomenclatures
                for size in nomenclature.sizes
                for sku in size.skus
            }

        for skus_group in track(
                sequence=grouper(skus, n=1000),
                description='Загрузка остатков по skus',
                console=console,
        ):
            for warehouse in warehouses:
                stocks_by_skus = wildberries_api_service.get_stocks_by_skus(warehouse_id=warehouse.id,
                                                                            skus=skus_group)
                if stocks_by_skus:
                    warehouse_id_to_stocks_by_skus[warehouse.id] += stocks_by_skus
                    break

    warehouses_stocks = [
        models.WarehouseStocks(warehouse_id=warehouse_id, stocks=stocks_by_skus)
        for warehouse_id, stocks_by_skus in warehouse_id_to_stocks_by_skus.items()
    ]
    generate_stocks_report_file(file_path=file_path, warehouses_stocks=warehouses_stocks)
    console.log('Остатки выгружены')


def download_prices(console: Console, api_key: str, file_path: str | pathlib.Path):
    category_to_nomenclature_prices: DefaultDict[str, list[models.NomenclaturePrice]] = collections.defaultdict(list)

    with closing_wildberries_api_http_client(api_key=api_key) as http_client:
        wildberries_api_service = WildberriesAPIService(http_client)

        with console.status('Загрузка цен...', spinner='bouncingBall'):
            nomenclature_prices = wildberries_api_service.get_prices(models.QuantityStatus.ANY)

        nomenclature_id_to_nomenclature_price: dict[int, models.NomenclaturePrice] = {
            nomenclature_price.nomenclature_id: nomenclature_price
            for nomenclature_price in nomenclature_prices
        }

        for nomenclatures in track(wildberries_api_service.get_nomenclatures()):

            for nomenclature in nomenclatures:
                nomenclature_price = nomenclature_id_to_nomenclature_price.get(nomenclature.id)
                if nomenclature_price is None:
                    continue
                category_to_nomenclature_prices[nomenclature.object].append(nomenclature_price)

    categories_prices = [
        models.CategoryPrices(
            category_name=category,
            nomenclature_prices=nomenclature_prices,
        ) for category, nomenclature_prices in category_to_nomenclature_prices.items()
    ]

    generate_prices_report_file(file_path=file_path, categories_prices=categories_prices)
    console.log('Цены выгружены')


def validate_excel_file(console: Console, file_path: pathlib.Path):
    try:
        workbook: Workbook = openpyxl.load_workbook(file_path)
    except InvalidFileException:
        console.print(f'[b][u]{file_path.name}[/u][/b] не является excel файлом')
    else:
        workbook.close()


def terminate(console: Console):
    Prompt.ask('Goodbye', console=console)
    exit(0)


def require_api_key(console: Console) -> str:
    while True:
        api_key = Prompt.ask('Введите API ключ')
        try:
            validate_api_key(api_key)
        except exceptions.ValidationError as error:
            console.print(str(error))
        else:
            return api_key


def require_file_not_exists(console: Console, file_path: pathlib.Path):
    if file_path.exists():
        is_replace_template_confirmed = Confirm.ask('Шаблон уже присутствует в директории. Заменить?')
        if not is_replace_template_confirmed:
            terminate(console)


def require_file(console: Console, file_path: pathlib.Path):
    if not file_path.exists():
        console.print(f'Отсутствует файл [b][u]{file_path.name}[/u][/b] в текущей директории')
        terminate(console)


def run(console: Console):
    action_number = Prompt.ask(
        prompt=(
            'Выберите действие'
            '\n1 - Загрузить остатки'
            '\n2 - Загрузить цены'
            '\n3 - Выгрузить все остатки'
            '\n4 - Выгрузить все цены'
            '\n5 - Сгенерировать шаблон'
            '\n0 - Выход\n'
        ),
        choices=['1', '2', '3', '4', '5', '0'],
    )
    match int(action_number):
        case 0:
            terminate(console)
        case 1:
            file_path = pathlib.Path('./остатки.xlsx')
            require_file(console, file_path)
            api_key = require_api_key(console)
            upload_stocks(console, api_key, file_path)
        case 2:
            file_path = pathlib.Path('./цены.xlsx')
            require_file(console, file_path)
            api_key = require_api_key(console)
            upload_prices(console, api_key, file_path)
        case 3:
            file_path = pathlib.Path('./выгрузка остатков.xlsx')
            require_file_not_exists(console, file_path)
            api_key = require_api_key(console)
            download_stocks(console, api_key, file_path)
        case 4:
            file_path = pathlib.Path('./выгрузка цен.xlsx')
            require_file_not_exists(console, file_path)
            api_key = require_api_key(console)
            download_prices(console, api_key, file_path)
        case 5:
            file_path = pathlib.Path('./шаблон.xlsx')
            require_file_not_exists(console, file_path)
            generate_template_file(file_path)
            console.log('Шаблон сохранен')
    terminate(console)
