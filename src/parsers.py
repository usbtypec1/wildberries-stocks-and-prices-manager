from typing import TypeAlias

from openpyxl.cell import ReadOnlyCell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

import models

PricesWorksheetRow: TypeAlias = tuple[ReadOnlyCell, ReadOnlyCell, ReadOnlyCell]
StocksWorksheetRow: TypeAlias = tuple[ReadOnlyCell, ReadOnlyCell, ReadOnlyCell, ReadOnlyCell]


class WorkbookParser:

    def __init__(self, workbook: Workbook):
        self.__workbook = workbook

    def parse_api_keys_worksheet(self) -> models.ParsedSupplierCredentialsWorksheet:
        worksheet: Worksheet = self.__workbook['Ключи']
        rows: tuple[tuple[ReadOnlyCell, ReadOnlyCell], ...] = worksheet['A2': F'B{worksheet.max_row}']

        parsed_rows: list[models.SupplierCredentialsWorksheetRow] = []
        error_row_numbers: list[int] = []

        for row_number, row in enumerate(rows, start=2):
            shop_name, api_key = [cell.value for cell in row]

            try:
                parsed_row = models.SupplierCredentialsWorksheetRow(shop_name=shop_name, api_key=api_key)
            except ValidationError:
                error_row_numbers.append(row_number)
            else:
                parsed_rows.append(parsed_row)

        return models.ParsedSupplierCredentialsWorksheet(rows=parsed_rows, error_row_numbers=error_row_numbers)

    def parse_stocks_worksheet(self) -> models.ParsedStocksWorksheet:
        worksheet: Worksheet = self.__workbook['Остатки']
        rows: tuple[StocksWorksheetRow, ...] = worksheet['A2': f'D{worksheet.max_row}']

        parsed_rows: list[models.StocksWorksheetRow] = []
        error_row_numbers: list[int] = []

        for row_number, row in enumerate(rows, start=2):
            shop_name, warehouse_id, sku, stocks_amount = [cell.value for cell in row]

            try:
                parsed_row = models.StocksWorksheetRow(
                    shop_name=shop_name,
                    warehouse_id=warehouse_id,
                    sku=sku,
                    stocks_amount=stocks_amount
                )
            except ValidationError:
                error_row_numbers.append(row_number)
            else:
                parsed_rows.append(parsed_row)

        return models.ParsedStocksWorksheet(rows=parsed_rows, error_row_numbers=error_row_numbers)

    # def parse_prices_worksheet(self) -> list[models.SupplierNomenclaturePricesToUpdate]:
    #     worksheet: Worksheet = self.__workbook['Цены']
    #     rows: tuple[PricesWorksheetRow, ...] = worksheet['A2': f'C{worksheet.max_row}']
    #     shop_name_to_nomenclatures = collections.defaultdict(set)
    #     for row_number, row in enumerate(rows, start=2):
    #         shop_name, nomenclature_id, price = [cell.value for cell in row]
    #         shop_name_to_nomenclatures[shop_name].add((row_number, nomenclature_id, price))
    #
    #     suppliers_nomenclatures = []
    #     for shop_name, nomenclatures in shop_name_to_nomenclatures.items():
    #
    #         parsed_nomenclatures = []
    #         for row_number, nomenclature_id, price in nomenclatures:
    #             try:
    #                 parsed_nomenclatures.append(models.NomenclaturePriceToUpdate(
    #                     nomenclature_id=nomenclature_id,
    #                     price=price,
    #                 ))
    #             except ValidationError:
    #                 print(f'Ошибка в строке №{row_number}')
    #
    #         suppliers_nomenclatures.append(models.SupplierNomenclaturePricesToUpdate(
    #             shop_name=shop_name,
    #             nomenclatures=parsed_nomenclatures,
    #         ))
    #     return suppliers_nomenclatures

# def match_price_nomenclatures_and_credentials(
#         suppliers_credentials: Iterable[models.SupplierCredentials],
#         suppliers_price_nomenclatures: Iterable[models.SupplierNomenclaturePricesToUpdate],
# ) -> list[models.SupplierNomenclaturePricesWithApiKeyToUpdate]:
#     shop_name_to_api_key = {credentials.shop_name: credentials.api_key for credentials in suppliers_credentials}
#
#     suppliers_price_nomenclatures_with_api_key: list[models.SupplierNomenclaturePricesWithApiKeyToUpdate] = []
#     missing_api_key_shop_names: set[str] = set()
#
#     for supplier_price_nomenclatures in suppliers_price_nomenclatures:
#         try:
#             api_key = shop_name_to_api_key[supplier_price_nomenclatures.shop_name]
#         except KeyError:
#             missing_api_key_shop_names.add(supplier_price_nomenclatures.shop_name)
#         else:
#             suppliers_price_nomenclatures_with_api_key.append(
#                 models.SupplierNomenclaturePricesWithApiKeyToUpdate(
#                     **supplier_price_nomenclatures.dict(),
#                     api_key=api_key,
#                 )
#             )
#
#     for shop_name in missing_api_key_shop_names:
#         print(f'Отсутствуют API ключи для магазина {shop_name}')
#
#     return suppliers_price_nomenclatures_with_api_key
#
