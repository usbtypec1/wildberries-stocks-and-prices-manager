import contextlib
import logging
import pathlib

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def generate_prices_template(filepath: str | pathlib.Path) -> None:
    workbook = openpyxl.Workbook()
    with contextlib.closing(workbook):
        prices_worksheet: Worksheet = workbook.active
        prices_worksheet.title = 'Цены'
        prices_worksheet['A1'].value = 'Название магазина'
        prices_worksheet['B1'].value = 'nm ID'
        prices_worksheet['C1'].value = 'Новая цена'
        prices_worksheet.column_dimensions['A'].width = 20
        prices_worksheet.column_dimensions['B'].width = 20
        prices_worksheet.column_dimensions['C'].width = 15

        stocks_worksheet: Worksheet = workbook.create_sheet('Остатки')
        stocks_worksheet['A1'].value = 'Название магазина'
        stocks_worksheet['B1'].value = 'ID склада'
        stocks_worksheet['C1'].value = 'Баркод'
        stocks_worksheet['D1'].value = 'Новое количество остатков'
        stocks_worksheet.column_dimensions['A'].width = 20
        stocks_worksheet.column_dimensions['B'].width = 20
        stocks_worksheet.column_dimensions['C'].width = 20
        stocks_worksheet.column_dimensions['D'].width = 22

        api_keys_worksheet: Worksheet = workbook.create_sheet('Ключи')
        api_keys_worksheet['A1'].value = 'Название магазина'
        api_keys_worksheet['B1'].value = 'API ключ'
        api_keys_worksheet.column_dimensions['A'].width = 20
        api_keys_worksheet.column_dimensions['B'].width = 40

        workbook.save(filepath)
    logging.debug(f'Template file was created in {filepath}')
