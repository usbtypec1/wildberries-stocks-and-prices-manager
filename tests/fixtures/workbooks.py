import contextlib
import pathlib

import pytest
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


@pytest.fixture
def workbook_only_with_valid_stocks_worksheet() -> Workbook:
    file_path = pathlib.Path.joinpath(
        pathlib.Path(__file__).parent.parent,
        'workbooks',
        'only-valid-stocks-worksheet.xlsx',
    )
    workbook = load_workbook(file_path, read_only=True)
    with contextlib.closing(workbook):
        yield workbook


@pytest.fixture
def workbook_with_invalid_stocks_worksheet() -> Workbook:
    file_path = pathlib.Path.joinpath(
        pathlib.Path(__file__).parent.parent,
        'workbooks',
        'stocks-worksheet-with-invalid-rows.xlsx',
    )
    workbook = load_workbook(file_path, read_only=True)
    with contextlib.closing(workbook):
        yield workbook


@pytest.fixture
def workbook_with_valid_api_keys_worksheet() -> Workbook:
    file_path = pathlib.Path.joinpath(
        pathlib.Path(__file__).parent.parent,
        'workbooks',
        'only-valid-api-keys-worksheet.xlsx',
    )
    workbook = load_workbook(file_path, read_only=True)
    with contextlib.closing(workbook):
        yield workbook


@pytest.fixture
def workbook_with_invalid_api_keys_worksheet() -> Workbook:
    file_path = pathlib.Path.joinpath(
        pathlib.Path(__file__).parent.parent,
        'workbooks',
        'api-keys-worksheet-with-invalid-rows.xlsx',
    )
    workbook = load_workbook(file_path, read_only=True)
    with contextlib.closing(workbook):
        yield workbook
